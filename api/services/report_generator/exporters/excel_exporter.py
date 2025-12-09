"""
Excel (XLSX) exporter for reports.
Creates professional Excel workbooks with multiple sheets and charts.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Dict, List, Any


class ExcelExporter:
    """Export reports to Excel format."""
    
    def __init__(self):
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
    
    def create_summary_sheet(self, data: Dict[str, Any]):
        """
        Create summary sheet with overview and charts.
        
        Args:
            data: Report data dictionary
        """
        ws = self.workbook.create_sheet("Summary")
        
        # Title
        ws['A1'] = data.get('title', 'Vulnerability Report')
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        # Generated date
        ws['A2'] = f"Generated: {data.get('generated_at', 'N/A')}"
        ws['A2'].font = Font(italic=True)
        
        # Summary statistics
        row = 4
        ws[f'A{row}'] = "Summary Statistics"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        stats = data.get('summary_stats', {})
        ws[f'A{row}'] = "Total Vulnerabilities:"
        ws[f'B{row}'] = stats.get('total_vulnerabilities', 0)
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Critical & High:"
        ws[f'B{row}'] = stats.get('critical_high_count', 0)
        ws[f'B{row}'].font = Font(bold=True, color="FF0000")
        row += 1
        
        # Severity breakdown
        row += 1
        ws[f'A{row}'] = "Severity Breakdown"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        severity_counts = stats.get('severity_counts', {})
        severity_colors = {
            'Critical': 'C00000',
            'High': 'FF6600',
            'Medium': 'FFC000',
            'Low': '92D050',
            'Info': '00B0F0'
        }
        
        for severity, color in severity_colors.items():
            ws[f'A{row}'] = severity
            ws[f'B{row}'] = severity_counts.get(severity.lower(), 0)
            ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws[f'A{row}'].font = Font(color="FFFFFF", bold=True)
            row += 1
        
        # Add pie chart for severity distribution
        if any(severity_counts.values()):
            chart = PieChart()
            labels = Reference(ws, min_col=1, min_row=row-5, max_row=row-1)
            data_ref = Reference(ws, min_col=2, min_row=row-5, max_row=row-1)
            chart.add_data(data_ref)
            chart.set_categories(labels)
            chart.title = "Vulnerabilities by Severity"
            ws.add_chart(chart, f"D{row-8}")
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def create_vulnerabilities_sheet(self, vulnerabilities: List[Dict[str, Any]]):
        """
        Create vulnerabilities sheet with detailed list.
        
        Args:
            vulnerabilities: List of vulnerability dictionaries
        """
        ws = self.workbook.create_sheet("Vulnerabilities")
        
        # Headers
        headers = ['ID', 'Name', 'Severity', 'CVSS', 'CVE', 'Target', 'Port', 'Status', 'Discovered']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        severity_colors = {
            'critical': 'C00000',
            'high': 'FF6600',
            'medium': 'FFC000',
            'low': '92D050',
            'info': '00B0F0'
        }
        
        for row, vuln in enumerate(vulnerabilities, 2):
            ws.cell(row=row, column=1, value=vuln.get('id'))
            ws.cell(row=row, column=2, value=vuln.get('name'))
            
            severity_cell = ws.cell(row=row, column=3, value=vuln.get('severity', '').upper())
            severity = vuln.get('severity', '').lower()
            if severity in severity_colors:
                severity_cell.fill = PatternFill(
                    start_color=severity_colors[severity],
                    end_color=severity_colors[severity],
                    fill_type="solid"
                )
                severity_cell.font = Font(color="FFFFFF", bold=True)
            
            ws.cell(row=row, column=4, value=vuln.get('cvss_score'))
            ws.cell(row=row, column=5, value=vuln.get('cve_id', 'N/A'))
            ws.cell(row=row, column=6, value=vuln.get('target'))
            ws.cell(row=row, column=7, value=vuln.get('port'))
            ws.cell(row=row, column=8, value=vuln.get('status', 'Open'))
            ws.cell(row=row, column=9, value=vuln.get('discovered_at'))
        
        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['B'].width = 40  # Name column wider
    
    def create_recommendations_sheet(self, recommendations: List[Dict[str, Any]]):
        """
        Create recommendations sheet.
        
        Args:
            recommendations: List of recommendation dictionaries
        """
        ws = self.workbook.create_sheet("Recommendations")
        
        # Headers
        ws['A1'] = "Priority Recommendations"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        
        # Headers for table
        headers = ['Priority', 'Vulnerability', 'Recommendation']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Data
        for row, rec in enumerate(recommendations, 4):
            ws.cell(row=row, column=1, value=rec.get('priority', 'Medium'))
            ws.cell(row=row, column=2, value=rec.get('vulnerability'))
            ws.cell(row=row, column=3, value=rec.get('recommendation'))
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 60
    
    def export(self, data: Dict[str, Any]) -> BytesIO:
        """
        Export data to Excel format.
        
        Args:
            data: Complete report data
            
        Returns:
            BytesIO object containing Excel file
        """
        # Create sheets
        self.create_summary_sheet(data)
        
        if 'vulnerabilities' in data:
            self.create_vulnerabilities_sheet(data['vulnerabilities'])
        
        if 'recommendations' in data:
            self.create_recommendations_sheet(data['recommendations'])
        
        # Save to BytesIO
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        return output
